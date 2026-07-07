import json
import os
import unicodedata
from datetime import datetime, date
from flask import render_template, request, redirect, url_for, session, flash, jsonify
from db import get_db
from utils import login_required, admin_required, calcular_edad, get_user_info, hoy
from itsdangerous import URLSafeSerializer, BadSignature
from werkzeug.security import generate_password_hash, check_password_hash

def register_routes(app):
    @app.route("/", methods=["GET", "POST"])
    def login():
        if "usuario" in session:
            return redirect(url_for("dashboard"))

        if request.method == "POST":
            identificacion = request.form["identificacion"].strip()
            contrasena = request.form["contrasena"].strip()

            if not identificacion or not contrasena:
                flash("Todos los campos son obligatorios.", "error")
                return render_template("login.html")

            conn = get_db()
            user = conn.execute(
                "SELECT * FROM usuarios WHERE identificacion = ?", (identificacion,)
            ).fetchone()

            if user:
                if not check_password_hash(user["contrasena"], contrasena):
                    flash("Credenciales incorrectas.", "error")
                    conn.close()
                    return render_template("login.html")
                
                if not user["activo"]:
                    flash("Su usuario se encuentra deshabilitado. Por favor contacte al administrador.", "error")
                    conn.close()
                    return render_template("login.html")
                
                if user.get("fecha_validez"):
                    from datetime import date
                    validez = user["fecha_validez"]
                    if isinstance(validez, str):
                        try:
                            validez = datetime.strptime(validez, "%Y-%m-%d").date()
                        except ValueError:
                            pass
                    if isinstance(validez, (date, datetime)):
                        if isinstance(validez, datetime):
                            validez = validez.date()
                        if hoy() > validez:
                            flash("Su cuenta ha expirado por fecha de validez. Por favor contacte al administrador.", "error")
                            conn.close()
                            return render_template("login.html")
            else:
                flash("Usuario no registrado. Comuníquese con el administrador para registrarse.", "error")
                conn.close()
                return render_template("login.html")

            conn.close()

            try:
                formularios_acceso_raw = user["formularios_acceso"]
                formularios_acceso = json.loads(formularios_acceso_raw) if formularios_acceso_raw else []
            except Exception:
                formularios_acceso = []

            # Parse perfiles (JSON array)
            try:
                perfiles_raw = user["perfil"]
                perfiles = json.loads(perfiles_raw) if perfiles_raw else []
                if isinstance(perfiles, str):
                    perfiles = [perfiles]
            except (json.JSONDecodeError, TypeError):
                perfiles = [user["perfil"]] if user["perfil"] else []

            # If multiple profiles, redirect to profile selection
            if len(perfiles) > 1:
                session["pendiente_seleccion_perfil"] = {
                    "nombre": user["nombre"],
                    "identificacion": user["identificacion"],
                    "registro_medico": user["registro_medico"],
                    "rol": user["rol"],
                    "perfiles": perfiles,
                    "requiere_cambio_clave": user["requiere_cambio_clave"],
                    "formularios_acceso": formularios_acceso,
                }
                return redirect(url_for("seleccionar_perfil"))

            # Single profile (or none): proceed normally
            perfil_unico = perfiles[0] if perfiles else ""
            
            # Extract profile-specific form access
            acc_dict = formularios_acceso
            if isinstance(acc_dict, str):
                try:
                    acc_dict = json.loads(acc_dict)
                except:
                    acc_dict = []
                    
            if isinstance(acc_dict, dict):
                formularios_acceso_list = acc_dict.get(perfil_unico, [])
            elif isinstance(acc_dict, list):
                formularios_acceso_list = acc_dict
            else:
                formularios_acceso_list = []

            session["usuario"] = {
                "nombre": user["nombre"],
                "identificacion": user["identificacion"],
                "registro_medico": user["registro_medico"],
                "rol": "admin" if (user["rol"] == "admin" or perfil_unico == "Administrador") else "usuario",
                "perfil": perfil_unico,
                "requiere_cambio_clave": user["requiere_cambio_clave"],
                "formularios_acceso": formularios_acceso_list,
            }
            session.permanent = True
            
            if user["requiere_cambio_clave"] == 1:
                flash("Debe cambiar su contraseña por seguridad al ser el primer ingreso o tras un restablecimiento.", "error")
                return redirect(url_for("cambiar_contrasena"))

            session["mostrar_bienvenida"] = True
            return redirect(url_for("dashboard"))

        return render_template("login.html")


    @app.route("/seleccionar_perfil", methods=["GET", "POST"])
    def seleccionar_perfil():
        pendiente = session.get("pendiente_seleccion_perfil")
        if not pendiente:
            return redirect(url_for("login"))

        if request.method == "POST":
            perfil_elegido = request.form.get("perfil", "").strip()
            perfiles_disponibles = pendiente.get("perfiles", [])

            if perfil_elegido not in perfiles_disponibles:
                flash("Perfil no válido. Seleccione uno de los perfiles asignados.", "error")
                return render_template("seleccionar_perfil.html", perfiles=perfiles_disponibles, nombre=pendiente["nombre"])

            # Extract profile-specific form access
            acc_dict = pendiente["formularios_acceso"]
            if isinstance(acc_dict, str):
                try:
                    acc_dict = json.loads(acc_dict)
                except:
                    acc_dict = []
                    
            if isinstance(acc_dict, dict):
                formularios_acceso_list = acc_dict.get(perfil_elegido, [])
            elif isinstance(acc_dict, list):
                formularios_acceso_list = acc_dict
            else:
                formularios_acceso_list = []

            # Set definitive session
            session["usuario"] = {
                "nombre": pendiente["nombre"],
                "identificacion": pendiente["identificacion"],
                "registro_medico": pendiente["registro_medico"],
                "rol": "admin" if (pendiente["rol"] == "admin" or perfil_elegido == "Administrador") else "usuario",
                "perfil": perfil_elegido,
                "requiere_cambio_clave": pendiente["requiere_cambio_clave"],
                "formularios_acceso": formularios_acceso_list,
            }
            session.permanent = True
            # Clean up temp data
            session.pop("pendiente_seleccion_perfil", None)

            if pendiente["requiere_cambio_clave"] == 1:
                flash("Debe cambiar su contraseña por seguridad al ser el primer ingreso o tras un restablecimiento.", "error")
                return redirect(url_for("cambiar_contrasena"))

            session["mostrar_bienvenida"] = True
            return redirect(url_for("dashboard"))

        return render_template("seleccionar_perfil.html", perfiles=pendiente["perfiles"], nombre=pendiente["nombre"])


    @app.route("/dashboard")
    @login_required
    def dashboard():
        user_ident = session["usuario"]["identificacion"]
        user_perfil = session["usuario"].get("perfil", "")
        conn = get_db()
        
        # 1. Historia Clínica Individual (pacientes where atencion_colectiva_id IS NULL and finalizado = 0)
        hc_pending = conn.execute(
            "SELECT COUNT(*) as count FROM pacientes WHERE atencion_colectiva_id IS NULL AND finalizado = 0 AND registrado_por_identificacion = ? AND perfil_registrador = ?",
            (user_ident, user_perfil)
        ).fetchone()["count"]
        
        # 2. Atención Colectiva (atencion_colectiva where finalizado = 0)
        ac_pending = conn.execute(
            "SELECT COUNT(*) as count FROM atencion_colectiva WHERE finalizado = 0 AND registrado_por_identificacion = ? AND perfil_registrador = ?",
            (user_ident, user_perfil)
        ).fetchone()["count"]
        
        # 3. Preoperacional (preoperacional where finalizado = 0)
        try:
            preop_pending = conn.execute(
                "SELECT COUNT(*) as count FROM preoperacional WHERE finalizado = 0 AND registrado_por_identificacion = ? AND perfil_registrador = ?",
                (user_ident, user_perfil)
            ).fetchone()["count"]
        except Exception:
            preop_pending = 0

        # 4. Checklists (tam, tab, pasb, pasm, equipos where finalizado = 0)
        checklist_pending = 0
        for _cl_table in ("checklist_tam", "checklist_tab", "checklist_pasb", "checklist_pasm", "checklist_equipos", "checklist_avanzada"):
            try:
                n = conn.execute(
                    f"SELECT COUNT(*) as count FROM {_cl_table} WHERE finalizado = 0 AND registrado_por_identificacion = ? AND perfil_registrador = ?",
                    (user_ident, user_perfil)
                ).fetchone()["count"]
                checklist_pending += n
            except Exception:
                pass

        total_pending = hc_pending + ac_pending
        checklist_total_pending = checklist_pending + preop_pending

        vehiculo_alerts = []
        if session["usuario"].get("rol") == "admin":
            try:
                active_vehiculos = conn.execute("SELECT id, placa, soat_vigencia, rtm_vigencia FROM vehiculos WHERE activo = 1").fetchall()
                fecha_hoy = hoy()
                for v in active_vehiculos:
                    placa = v["placa"]
                    for doc_name, doc_date in [("SOAT", v["soat_vigencia"]), ("RTM", v["rtm_vigencia"])]:
                        if doc_date:
                            if isinstance(doc_date, str):
                                try:
                                    d_val = datetime.strptime(doc_date.split()[0], "%Y-%m-%d").date()
                                except Exception:
                                    d_val = None
                            else:
                                d_val = doc_date if isinstance(doc_date, date) else doc_date.date()
                            
                            if d_val:
                                diff = (d_val - fecha_hoy).days
                                if diff in (1, 2, 3, 4, 5, 10, 15):
                                    vehiculo_alerts.append({
                                        "placa": placa,
                                        "tipo": doc_name,
                                        "dias": diff,
                                        "fecha": d_val.strftime("%d/%m/%Y")
                                    })
            except Exception as e:
                print("Error al calcular alertas de vehiculos:", e)

        conn.close()
        
        # Ordenar las alertas por los dias mas cercanos primero
        vehiculo_alerts.sort(key=lambda x: x["dias"])

        return render_template(
            "dashboard.html",
            total_pending=total_pending,
            hc_pending=hc_pending,
            ac_pending=ac_pending,
            preop_pending=preop_pending,
            checklist_pending=checklist_pending,
            checklist_total_pending=checklist_total_pending,
            vehiculo_alerts=vehiculo_alerts
        )


    @app.route("/logout")
    def logout():
        if "usuario" in session:
            conn = get_db()
            conn.execute("""
                UPDATE usuarios 
                SET ultima_latitud = NULL, ultima_longitud = NULL, ultima_actualizacion_gps = NULL 
                WHERE identificacion = ?
            """, (session["usuario"]["identificacion"],))
            conn.commit()
            conn.close()
        session.clear()
        return redirect(url_for("login"))


    @app.route("/cambiar_contrasena", methods=["GET", "POST"])
    def cambiar_contrasena():
        if "usuario" not in session:
            return redirect(url_for("login"))
            
        if request.method == "POST":
            nueva_clave = request.form["nueva_clave"].strip()
            confirmacion = request.form["confirmacion"].strip()

            if not nueva_clave or not confirmacion:
                flash("Todos los campos son obligatorios.", "error")
                return render_template("cambiar_contrasena.html")

            if nueva_clave != confirmacion:
                flash("Las contraseñas no coinciden.", "error")
                return render_template("cambiar_contrasena.html")

            identificacion = session["usuario"]["identificacion"]
            if nueva_clave == identificacion:
                flash("La nueva contraseña no puede ser igual a su número de identificación por seguridad.", "error")
                return render_template("cambiar_contrasena.html")

            conn = get_db()
            hashed_clave = generate_password_hash(nueva_clave)
            conn.execute(
                "UPDATE usuarios SET contrasena = ?, requiere_cambio_clave = 0 WHERE identificacion = ?",
                (hashed_clave, identificacion)
            )
            conn.commit()
            conn.close()

            session["usuario"]["requiere_cambio_clave"] = 0
            flash("Contraseña actualizada exitosamente.", "success")
            return redirect(url_for("dashboard"))

        return render_template("cambiar_contrasena.html")


    @app.route("/recuperar_contrasena", methods=["GET", "POST"])
    def recuperar_contrasena():
        if request.method == "POST":
            identificacion = request.form.get("identificacion", "").strip()
            if not identificacion:
                flash("Debe ingresar su identificación.", "error")
            else:
                conn = get_db()
                user = conn.execute("SELECT * FROM usuarios WHERE identificacion = ?", (identificacion,)).fetchone()
                
                if user and user.get("correo"):
                    import random
                    import string
                    temp_password = ''.join(random.choices(string.digits, k=4))
                    
                    hashed_temp = generate_password_hash(temp_password)
                    conn.execute("UPDATE usuarios SET contrasena = ?, requiere_cambio_clave = 1 WHERE id = ?", (hashed_temp, user["id"]))
                    conn.commit()
                    
                    from utils import send_recovery_email
                    success = send_recovery_email(user["correo"], temp_password)
                    if success:
                        flash("Si la identificación existe en nuestro sistema, se han enviado las instrucciones al correo registrado.", "success")
                    else:
                        flash("Ocurrió un error al enviar el correo. Por favor contacte al administrador.", "error")
                else:
                    # Prevent user enumeration by flashing the same message
                    flash("Si la identificación existe en nuestro sistema, se han enviado las instrucciones al correo registrado.", "success")
                conn.close()
            return redirect(url_for("login"))
        return render_template("recuperar_contrasena.html")


    @app.route("/usuarios/plantilla_excel")
    @login_required
    @admin_required
    def plantilla_excel():
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Usuarios"

        headers = [
            "nombre", "identificacion", "registro_medico", "correo",
            "rol", "perfiles", "activo"
        ]
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="1e6fbf")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        ws.cell(row=2, column=1, value="Ejemplo Pérez")
        ws.cell(row=2, column=2, value="1234567890")
        ws.cell(row=2, column=3, value="RM-12345")
        ws.cell(row=2, column=4, value="ejemplo@correo.com")
        ws.cell(row=2, column=5, value="usuario")
        ws.cell(row=2, column=6, value="Médico,APH")
        ws.cell(row=2, column=7, value=1)

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 22

        from io import BytesIO
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        from flask import send_file
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name="plantilla_usuarios.xlsx")


    @app.route("/usuarios/importar_excel", methods=["POST"])
    @login_required
    @admin_required
    def importar_usuarios_excel():
        import openpyxl
        from openpyxl.utils import get_column_letter

        file = request.files.get("archivo_excel")
        if not file or file.filename == "":
            flash("Debe seleccionar un archivo Excel.", "error")
            return redirect(url_for("usuarios"))

        if not file.filename.lower().endswith((".xlsx", ".xls")):
            flash("El archivo debe tener extensión .xlsx o .xls.", "error")
            return redirect(url_for("usuarios"))

        conn = get_db()
        creados = 0
        creados_list = []
        errores = []
        fila = 1

        try:
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                flash("El archivo Excel está vacío.", "error")
                conn.close()
                return redirect(url_for("usuarios"))

            header_row = [str(c).strip().lower() if c else "" for c in rows[0]]
            expected = ["nombre", "identificacion", "registro_medico", "correo", "rol", "perfiles", "activo"]
            col_map = {}
            for i, h in enumerate(expected):
                try:
                    col_map[h] = header_row.index(h)
                except ValueError:
                    errores.append(f"Fila 1: falta la columna '{h}'")
                    conn.close()
                    flash("El archivo no tiene las columnas requeridas. Descargue la plantilla.", "error")
                    return redirect(url_for("usuarios"))

            for row in rows[1:]:
                fila += 1
                if all(c is None or str(c).strip() == "" for c in row):
                    continue

                try:
                    nombre = str(row[col_map["nombre"]]).strip() if row[col_map["nombre"]] is not None else ""
                    identificacion = str(row[col_map["identificacion"]]).strip() if row[col_map["identificacion"]] is not None else ""
                    registro_medico = str(row[col_map["registro_medico"]]).strip() if row[col_map["registro_medico"]] is not None else ""
                    correo = str(row[col_map["correo"]]).strip() if row[col_map["correo"]] is not None else ""
                    rol = str(row[col_map["rol"]]).strip().lower() if row[col_map["rol"]] is not None else "usuario"
                    perfiles_raw = str(row[col_map["perfiles"]]).strip() if row[col_map["perfiles"]] is not None else ""
                    activo_raw = row[col_map["activo"]]
                except IndexError:
                    errores.append(f"Fila {fila}: datos incompletos")
                    continue

                if not nombre or not identificacion:
                    errores.append(f"Fila {fila}: nombre e identificación son obligatorios")
                    continue

                raw_perfiles = [p.strip() for p in perfiles_raw.split(",") if p.strip()]
                if not raw_perfiles:
                    errores.append(f"Fila {fila} ({nombre}): debe tener al menos un perfil")
                    continue

                def normalizar(s):
                    return unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode().strip().lower()

                perfiles_canon = {"Médico", "Enfermero", "APH", "Auxiliar en Enfermeria", "Socorrista", "Conductor"}
                mapa_norm = {normalizar(p): p for p in perfiles_canon}
                perfiles_list = []
                invalidos = []
                for p in raw_perfiles:
                    norm = normalizar(p)
                    if norm in mapa_norm:
                        perfiles_list.append(mapa_norm[norm])
                    else:
                        invalidos.append(p)
                if invalidos:
                    errores.append(f"Fila {fila} ({nombre}): perfil(es) inválido(s): {', '.join(invalidos)}. Válidos: {', '.join(sorted(perfiles_canon))}")
                    continue

                tiene_perfil_requerido = any(p in {"Médico", "Enfermero", "APH", "Auxiliar en Enfermeria"} for p in perfiles_list)
                if tiene_perfil_requerido and not registro_medico:
                    errores.append(f"Fila {fila} ({nombre}): la Resolución en Salud es obligatoria para los perfiles Médico, Enfermero, APH y Auxiliar en Enfermeria")
                    continue

                if rol not in ("usuario", "admin"):
                    rol = "usuario"

                try:
                    activo = int(activo_raw) if activo_raw is not None else 1
                except (ValueError, TypeError):
                    activo = 1

                existing = conn.execute("SELECT id FROM usuarios WHERE identificacion = ?", (identificacion,)).fetchone()
                if existing:
                    errores.append(f"Fila {fila} ({nombre}): identificación ya existe")
                    continue

                perfiles_json = json.dumps(perfiles_list, ensure_ascii=False)
                formularios_acceso_default = {}
                for p in perfiles_list:
                    formularios_acceso_default[p] = []
                formularios_acceso_json = json.dumps(formularios_acceso_default, ensure_ascii=False)
                hashed_pw = generate_password_hash(identificacion)

                conn.execute("""
                    INSERT INTO usuarios (nombre, identificacion, registro_medico, rol, perfil, activo, firma, contrasena, requiere_cambio_clave, formularios_acceso, correo)
                    VALUES (?, ?, ?, ?, ?, ?, '', ?, 1, ?, ?)
                """, (nombre, identificacion, registro_medico, rol, perfiles_json, activo, hashed_pw, formularios_acceso_json, correo))
                creados += 1
                creados_list.append({"nombre": nombre, "identificacion": identificacion})

            conn.commit()
        except Exception as e:
            conn.close()
            flash(f"Error al procesar el archivo: {e}", "error")
            return redirect(url_for("usuarios"))

        conn.close()

        session['import_result'] = {
            'creados': creados,
            'creados_list': creados_list,
            'errores': errores,
            'total_rows': len(rows) - 1 if len(rows) > 1 else 0
        }
        if creados > 0:
            plural = "" if creados == 1 else "s"
            flash(f"✅ {creados} usuario{plural} creado{plural} exitosamente.", "success")
        return redirect(url_for("usuarios"))


    @app.route("/usuarios/clear_import_result")
    @login_required
    @admin_required
    def clear_import_result():
        session.pop('import_result', None)
        return 'ok'


    @app.route("/usuarios", methods=["GET", "POST"])
    @login_required
    @admin_required
    def usuarios():
        conn = get_db()
        if request.method == "POST":
            nombre = request.form["nombre"].strip()
            identificacion = request.form["identificacion"].strip()
            registro_medico = request.form["registro_medico"].strip()
            rol = request.form["rol"].strip()
            perfiles_list = request.form.getlist("perfil")
            perfil = json.dumps(perfiles_list, ensure_ascii=False)
            activo = int(request.form.get("activo", 1))
            firma = request.form.get("firma", "").strip()
            correo = request.form.get("correo", "").strip()
            
            formularios_acceso_dict = {}
            for p in perfiles_list:
                formularios_acceso_dict[p] = request.form.getlist(f"formularios_acceso_{p}")
            formularios_acceso = json.dumps(formularios_acceso_dict, ensure_ascii=False)

            perfiles_requieren_rm = {"Médico", "Enfermero", "APH", "Auxiliar en Enfermeria"}
            tiene_perfil_requerido = any(p in perfiles_requieren_rm for p in perfiles_list)

            if not nombre or not identificacion:
                flash("Nombre e identificación son obligatorios.", "error")
            elif not perfiles_list:
                flash("Debe seleccionar al menos un perfil profesional.", "error")
            elif tiene_perfil_requerido and not registro_medico:
                flash("La Resolución en Salud es obligatoria para los perfiles Médico, Enfermero, APH y Auxiliar en Enfermeria.", "error")
            else:
                existing = conn.execute(
                    "SELECT * FROM usuarios WHERE identificacion = ?", (identificacion,)
                ).fetchone()
                if existing:
                    flash("Ya existe un usuario con esa Identificación.", "error")
                else:
                    # Password defaults to identification and requires change (requiere_cambio_clave = 1)
                    hashed_identificacion = generate_password_hash(identificacion)
                    fecha_validez = request.form.get("fecha_validez", "").strip() or None
                    conn.execute("""
                        INSERT INTO usuarios (nombre, identificacion, registro_medico, rol, perfil, activo, firma, contrasena, requiere_cambio_clave, formularios_acceso, correo, fecha_validez)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?, ?)
                    """, (nombre, identificacion, registro_medico, rol, perfil, activo, firma, hashed_identificacion, formularios_acceso, correo, fecha_validez))
                    conn.commit()
                    flash("Usuario creado exitosamente. La contraseña inicial es su número de identificación.", "success")
                    conn.close()
                    return redirect(url_for("usuarios"))

        usuarios_list = conn.execute("SELECT * FROM usuarios ORDER BY id DESC").fetchall()
        conn.close()
        return render_template("usuarios.html", usuarios=usuarios_list, usuario=session["usuario"], json=json)


    @app.route("/usuarios/editar/<int:user_id>", methods=["GET", "POST"])
    @login_required
    @admin_required
    def editar_usuario(user_id):
        conn = get_db()
        user = conn.execute("SELECT * FROM usuarios WHERE id = ?", (user_id,)).fetchone()
        if not user:
            flash("Usuario no encontrado.", "error")
            conn.close()
            return redirect(url_for("usuarios"))

        if request.method == "POST":
            nombre = request.form["nombre"].strip()
            identificacion = request.form["identificacion"].strip()
            registro_medico = request.form["registro_medico"].strip()
            rol = request.form["rol"].strip()
            perfiles_list = request.form.getlist("perfil")
            perfil = json.dumps(perfiles_list, ensure_ascii=False)
            activo = int(request.form.get("activo", 1))
            firma = request.form.get("firma", "").strip()
            correo = request.form.get("correo", "").strip()
            
            formularios_acceso_dict = {}
            for p in perfiles_list:
                formularios_acceso_dict[p] = request.form.getlist(f"formularios_acceso_{p}")
            formularios_acceso = json.dumps(formularios_acceso_dict, ensure_ascii=False)

            perfiles_requieren_rm = {"Médico", "Enfermero", "APH", "Auxiliar en Enfermeria"}
            tiene_perfil_requerido = any(p in perfiles_requieren_rm for p in perfiles_list)

            if not nombre or not identificacion:
                flash("Nombre e identificación son obligatorios.", "error")
            elif not perfiles_list:
                flash("Debe seleccionar al menos un perfil profesional.", "error")
            elif tiene_perfil_requerido and not registro_medico:
                flash("La Resolución en Salud es obligatoria para los perfiles Médico, Enfermero, APH y Auxiliar en Enfermeria.", "error")
            else:
                existing = conn.execute(
                    "SELECT * FROM usuarios WHERE identificacion = ? AND id != ?", (identificacion, user_id)
                ).fetchone()
                if existing:
                    flash("Ya existe otro usuario con esa Identificación.", "error")
                else:
                    fecha_validez = request.form.get("fecha_validez", "").strip() or None
                    conn.execute("""
                        UPDATE usuarios
                        SET nombre = ?, identificacion = ?, registro_medico = ?, rol = ?, perfil = ?, activo = ?, firma = ?, formularios_acceso = ?, correo = ?, fecha_validez = ?
                        WHERE id = ?
                    """, (nombre, identificacion, registro_medico, rol, perfil, activo, firma, formularios_acceso, correo, fecha_validez, user_id))
                    conn.commit()
                    flash("Usuario actualizado exitosamente.", "success")
                    conn.close()
                    return redirect(url_for("usuarios"))

        conn.close()
        return render_template("editar_usuario.html", user=user, usuario=session["usuario"], json=json)


    @app.route("/usuarios/toggle/<int:user_id>")
    @login_required
    @admin_required
    def toggle_usuario(user_id):
        conn = get_db()
        user = conn.execute("SELECT * FROM usuarios WHERE id = ?", (user_id,)).fetchone()
        if user:
            if user["identificacion"] == "admin":
                flash("No se puede deshabilitar al usuario administrador principal.", "error")
            else:
                nuevo_estado = 0 if user["activo"] else 1
                conn.execute("UPDATE usuarios SET activo = ? WHERE id = ?", (nuevo_estado, user_id))
                conn.commit()
                flash(f"Usuario {'habilitado' if nuevo_estado else 'deshabilitado'} correctamente.", "success")
        else:
            flash("Usuario no encontrado.", "error")
        conn.close()
        return redirect(url_for("usuarios"))


    @app.route("/usuarios/reset/<int:user_id>")
    @login_required
    @admin_required
    def reset_usuario_contrasena(user_id):
        conn = get_db()
        user = conn.execute("SELECT * FROM usuarios WHERE id = ?", (user_id,)).fetchone()
        if user:
            if user["identificacion"] == "admin":
                flash("No se puede restablecer la contraseña del administrador principal desde aquí.", "error")
            else:
                # Reset password to identification and force change
                hashed_id = generate_password_hash(user["identificacion"])
                conn.execute(
                    "UPDATE usuarios SET contrasena = ?, requiere_cambio_clave = 1 WHERE id = ?",
                    (hashed_id, user_id)
                )
                conn.commit()
                flash(f"Contraseña de {user['nombre']} restablecida. La nueva clave es su identificación.", "success")
        else:
            flash("Usuario no encontrado.", "error")
        conn.close()
        return redirect(url_for("usuarios"))

